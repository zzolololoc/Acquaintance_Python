import openpyxl
from openpyxl.utils.cell import get_column_letter


# Создание новой ниги
wb1 = openpyxl.Workbook()

# или же существующий файл
wb_mame: str = "существующий_файл"
try:
    wb2 = openpyxl.load_workbook(f'{wb_mame}.xlsx')
except FileNotFoundError:
    print(f"Файл '{wb_mame}.xlsx' не найден")


# Берем активный лист и назначаем ему sheet для работы с sheet :)
sheet = wb1.active

# Название листа книги
sheet.title = "Лист1"

# Запись данных в ячейки
# Запись "Привет" в ячейку A1
sheet['A1'] = "Привет"
# Запись "Мир" в ячейку A2
sheet['A2'] = "Мир"

# Название листа
print(sheet.title)

# Содержимое ячеек
# Напрямую из ячейки по индексу [A1] -> и тп
print(f"\nЗначение ячейки A1: {sheet['A1'].value}")
print(f"Значение ячейки A2: {sheet['A2'].value}")

# Метод тыка циклом
print("\nСодержимое столбxца A:")
for row in range(1, 3):
    cell = sheet[f'A{row}']                 # Получение ячейки по ее адресу (A1, A2 и т.д.)
    print(f"Ячейка A{row}: {cell.value}")


# Добавление нового значения
sheet['C9'] = "Что-то интересное"

# Потом вывод значение ячейки C9
print(f"Значение ячейки C9: {sheet['C9'].value}")

# Вложенный цикл (ага ага, тот самый роу кол сеточкой :) )
print("Содержимое нескольких ячеек:")
for row in range(1, 10):
    for col in range(1, 10):

        col_letter = get_column_letter(col)
        cell = sheet[f'{col_letter}{row}']
        print(f"Ячейка {col_letter}{row}: {cell.value}")


# Сохранение в ДИРЕКТОРИЮ ЗАПУСКА СКРИПТА
wb1.save('новый_файл.xlsx') # НОВЫЙ
# wb2.save(f'{wb_mame}.xlsx') # Существущий